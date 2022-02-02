import os
import re
import time
import json
import pandas
import random
import logging
import openpyxl
from pathlib import Path
from copy import deepcopy

from batf.SutInit import Sut
from batf.Report import ReportGen
from batf.Common.LogAnalyzer import LogAnalyzer
from batf import MiscLib, SshLib, SutInit, core, SerialLib
from SPR4P.Config import SutConfig
from SPR4P.Config.PlatConfig import Key, Msg, BootOS
from SPR4P.BaseLib import SetUpLib, PlatMisc, BmcLib


class BaseLine:
    def __init__(self, xlsx_file):
        self.xlsx = xlsx_file
        self.head_value = "Value"
        self.head_variable = "VariableName"
        self.head_unitool = "UniTool"

    def load_sheet(self):
        work_book = openpyxl.load_workbook(self.xlsx)
        self.sheet = work_book.active
        return self.sheet

    def get_header_data(self) -> dict:
        """
        return dict like:
        { value: column3, variable: column5, unitool: column7 }
        """
        self.load_sheet()
        header_index = {}
        for col in range(1, self.sheet.max_column + 1, 1):
            cell_value = self.sheet.cell(row=1, column=col).value.strip()
            if cell_value in [self.head_value, self.head_variable, self.head_unitool]:
                header_index[cell_value] = self.sheet.cell(row=1, column=col).col_idx
        return header_index

    def get_valid_data(self) -> dict:
        """return dict of all unitool support items like:
        {
        variable1:
                { value1: 0, value2: 1 },
        variable2:
                { value1: 0, value2: 1 }
        }
        """
        all_data = {}
        header_index = self.get_header_data()
        strikeout = 0
        non_stand = {}
        for irow in range(1, self.sheet.max_row + 1, 1):
            unitool_support = self.sheet.cell(row=irow, column=header_index[self.head_unitool]).value
            if (not unitool_support) or (not unitool_support.strip().startswith("Y")):
                continue
            if all(self.sheet.cell(row=irow, column=co).font.strike for co in header_index.values()):
                variable = self.sheet.cell(row=irow, column=header_index[self.head_variable]).value
                logging.debug(f"Row {irow}:{variable} - is ignored due to strikeout")  # 忽略划了删除线的选项
                strikeout += 1
                continue
            variable_name = self.sheet.cell(row=irow, column=header_index[self.head_variable]).value
            value_list = self.sheet.cell(row=irow, column=header_index[self.head_value]).value.strip().split("\n")
            val_spliter = [":", "："]  # 支持中英文冒号
            for val in value_list:
                if any(spliter in val for spliter in val_spliter):
                    value_dict = dict([val.rsplit(sp, 1) for sp in val_spliter if sp in val])
                    if variable_name in all_data:
                        all_data[variable_name].update(value_dict)
                    else:
                        all_data[variable_name] = value_dict
                else:
                    non_stand[variable_name] = val
        logging.debug(f"Total {strikeout} rows are strikeout")
        for non_k, non_v in non_stand.items():
            logging.warning(f"{non_k} = \"{non_v}\"")  # 不带冒号的Value选项
        return all_data

    def value_in_range(self) -> dict:
        """
        如过基线给定的Value类型是一个区间，就测试最大值和最小值
        return dict: 去掉 Max/Min  maximum/minimum 此类选项的Step
        """
        data = self.get_valid_data()
        idata = deepcopy(data)
        for key, value in data.items():
            for val in value.keys():
                if val.strip().lower() == "step":
                    idata[key].pop(val)
        return idata

    def variable_in_range(self) -> dict:
        """
        如果基线给定的变量名是一个区间, 则将其实例化, 每个选项都单独列出来, BIOS基线整理后的原始数据字典
        list[0]~[3] 转为 list, list[1], list[2], list[3]
        list[0]~[3],[5] 转为 list, list[1], list[2], list[3], list[5]
        list[0]/[3]/[5] 转为 list, list[3]， list[5]
        return dict
        """
        data = self.value_in_range()
        idata = {}
        for key, value in data.items():
            if "[" not in key:
                idata[key] = value
                continue
            var_name = "".join(re.findall("(\w+)\[", key))
            var_index = re.findall("\[(\d+)]", key)
            if len(var_index) == 1 and ("[0]" in key):  # 去掉下表为[0]的索引
                idata[var_name] = value
            elif len(var_index) == 2 and ("~" in key):  # 正常的索引区间
                low, high = list(map(int, var_index))
                for i in range(low, high+1, 1):
                    if str(low) == "0":
                        idata[f"{var_name}"] = value
                        continue
                    idata[f"{var_name}[{i}]"] = value
            elif len(var_index) > 2:  # 非标准的索引区间
                range_index = re.findall("\[(\d+)]~\[(\d+)]", key)  # 区间部分
                if range_index:
                    low, high = list(map(int, range_index[0]))
                    for j in range(low, high+1, 1):
                        if str(low) == "0":
                            idata[f"{var_name}"] = value
                            continue
                        idata[f"{var_name}[{j}]"] = value
                other_index = re.findall("\[(\d+)]", re.sub("(\[\d+]~\[\d+])", "", key))  # 多个索引单独列出部分
                if other_index:
                    for o_index in other_index:
                        if str(o_index) == "0":
                            idata[f"{var_name}"] = value
                            continue
                        idata[f"{var_name}[{o_index}]"] = value
            else:  # 不带下标的索引
                idata[key] = value
        with open(os.path.join(SutConfig.Env.LOG_DIR, "origin_baseline.json"), "w", encoding="utf-8") as _json:
            json.dump(idata, _json, indent=4, ensure_ascii=False)
        return idata

    def make_test_data(self) -> dict:
        """
        生成字典 {变量名:[value1, value2...], 0x开头的转为十进制
        return dict like:
            {
                variable1:[value1, value2],
                variable2:[value1, value2],
            }
        """
        data = self.variable_in_range()
        idata = {}
        for var_name, values in data.items():
            for val_name, val_index in values.items():
                val_index = int(val_index, 16) if val_index.lower().startswith("0x") else int(val_index, 10)
                if var_name in idata:
                    idata[var_name].append(val_index)
                else:
                    idata[var_name] = [val_index]
        return idata

class BaseLineLoop(BaseLine):
    def __init__(self, baseline, unitool):
        super(BaseLineLoop, self).__init__(baseline)
        self.test_data = self.make_test_data()
        self.unitool = unitool
        self.batch_try_count = 10
        self.init_report()

    def init_report(self):
        report = {}
        for name, values in self.test_data.items():
            for val in values:
                report[f"{name}:{val}"] = {"default": None, "error": None, "write": None, "check": None, "retry": None, "summary": None}
        self.report = pandas.DataFrame().from_dict(report, orient="index")

    def read_default_value(self) -> dict:
        """
        通过unitool读取所有选项的当前默认value, 顺便检查变量名是否正确, 是否可读
        没有读到默认值的重试一次
        """
        default = self.unitool.read(*self.test_data)
        for df, val in default.items():
            if val is not None:
                continue
            retry_read = self.unitool.read(df)
            if retry_read is None:
                logging.error(f"Read {df} error after retry")
            else:
                default.update(retry_read)
        for _df_key, _df_val in default.items():
            index_name = f"{_df_key}:{_df_val}"
            self.report.loc[index_name, "default"] = "yes"
            if _df_val is None:
                self.report.loc[index_name, "error"] = "yes"
        return default

    def drop_default_value(self):
        """测试数据去掉通过 unitool 读取到的默认值, 如果无法读到默认值，说明变量异常，跳过测试"""
        for name, default in self.read_default_value().items():
            try:
                self.test_data[name].remove(int(default))
            except Exception:
                logging.warning(f"Default {name} = {default} not in values list, variable maybe integer")

    def testdata_generator(self):
        """生成器，每次调用时，临时生成数量为 {self.batch_try_count} 的测试数据字典，用于批量测试"""
        check_data = deepcopy(self.test_data)
        while check_data:
            test_count = 0
            temp_data = {}
            for name in self.test_data:
                if name not in check_data:  # 变量全部测完，跳过
                    continue
                if not check_data[name]:  # value全部测完，删除整个变量
                    check_data.pop(name)
                    continue
                value_choice = random.choice(check_data[name])
                temp_data[name] = value_choice
                check_data[name].remove(value_choice)  # 单个value测完后删除此value
                test_count += 1
                if test_count >= self.batch_try_count:
                    break
            yield temp_data

    def single_write(self, data):
        """输入一个或多个选项，一项一项单独修改,并记录write结果"""
        for sw_key, sw_val in data.items():
            index_name = f"{sw_key}:{sw_val}"
            try:
                write_reslut = self.unitool.write(**{sw_key: sw_val})
                if write_reslut:
                    self.report.loc[index_name, "write"] = "pass"
                else:
                    self.report.loc[index_name, "write"] = "fail"
            except Exception:
                self.report.loc[index_name, "write"] = "error"

    def batch_write(self, data: dict):
        """输入一个或多个选项，批量修改，修改失败或报错再尝试单个修改, 并记录write结果"""
        try:
            if self.unitool.write(**data):
                for bw_key, bw_val in data.items():
                    index_name = f"{bw_key}:{bw_val}"
                    self.report.loc[index_name, "write"] = "pass"
            else:
                logging.info("batch write fail, retry write one by one...")
                self.single_write(data)
        except Exception:
            logging.info("batch write error, retry write one by one...")
            self.single_write(data)


    def single_check(self, data):
        """输入一个或多个选项，一项一项单独检查,并记录check结果"""
        for sc_key, sc_val in data.items():
            index_name = f"{sc_key}:{sc_val}"
            try:
                check_reslut = self.unitool.check(**{sc_key: sc_val})
                assert (check_reslut is not None)
                if check_reslut:
                    self.report.loc[index_name, "check"] = "pass"
                else:
                    self.report.loc[index_name, "check"] = "fail"
            except Exception:
                self.report.loc[index_name, "check"] = "error"

    def batch_check(self, data: dict):
        """输入一个或多个选项，批量检查，检查失败或报错再尝试单个检查, 并记录check结果"""
        try:
            if self.unitool.check(**data):
                for bc_key, bc_value in data.items():
                    index_name = f"{bc_key}:{bc_value}"
                    self.report.loc[index_name, "check"] = "pass"
            else:
                logging.info("batch check fail, retry check one by one...")
                self.single_check(data)
        except Exception:
            logging.info("batch check error, retry check one by one...")
            self.single_check(data)

    def report_summary(self, excel_save: str):
        self.report.sort_index(inplace=True)
        self.report["summary"] = self.report.apply(lambda x:
            "pass" if ((x.write == "pass" and x.check == "pass") or x.retry == "pass") else
            ("fail" if (x.write == "fail" or x.check == "fail" or x.retry == "fail") else None), axis=1)
        if excel_save.endswith(".xlsx"):
            self.report.to_excel(excel_save)


@core.test_case(("123", "[TC123] Uitool Variables Loop Test", "Uitool Variables Loop Test"))
def test_variable_loop():
    baseline = Path(__file__).parent.parent.absolute()/"Resource/SetupBase/V7服务器setup菜单基线版本_Byosoft_V0.1.xlsx"
    boot_delay = 1200
    reboot_count = 0

    def _reboot_ping(boottime=boot_delay):
        nonlocal reboot_count
        reboot_count += 1
        BmcLib.force_reset()
        time.sleep(30)
        if MiscLib.ping_sut(SutConfig.Env.OS_IP, boottime):
            return True
        if SetUpLib.reset_default():
            return
        BmcLib.clear_cmos()

    try:
        test = BaseLineLoop(baseline, unitool=Sut.UNITOOL)
        test.drop_default_value()

        for data in test.testdata_generator():  # 批量测试
            BmcLib.clear_cmos()  # 每组批量测试前clear CMOS
            _reboot_ping()

            test.batch_write(data)
            write_pass = {}  # 只有write成功的才去检查
            for data_k, data_v in data.items():
                if test.report.loc[f"{data_k}:{data_v}", "write"] == "pass":
                    write_pass[data_k] = data_v

            if not write_pass:  # 如果批量全部都设置失败，则开始下一组测试
                continue
            if _reboot_ping():
                test.batch_check(write_pass)
            else:
                for one_key, one_value in write_pass.items():  # write pass但重启无法开机的标为 boot_fail
                    test.report.loc[f"{one_key}:{one_value}", "check"] = "boot_fail"

        test.report_summary(os.path.join(SutConfig.Env.LOG_DIR, "unitool_batch_test.xlsx"))

        check_fail = test.report[test.report["check"].str.contains("fail", na=False)]  # 错误的一个一个修改加重启测试
        fail_count = len(check_fail.index)
        fcount = 0
        for _c_fail in check_fail.index:
            fcount += 1
            logging.info(f"Start to single retry {_c_fail} [{fcount}/{fail_count}]")  # 显示测试进度
            BmcLib.clear_cmos()
            _reboot_ping()
            try:
                _fail_key, _fail_value = _c_fail.rsplit(":", 1)
                if Sut.UNITOOL.write(**{_fail_key: _fail_value}):
                    if _reboot_ping():
                        if Sut.UNITOOL.check(**{_fail_value: _fail_value}):
                            test.report.loc[_c_fail, "retry"] = "pass"
                        else:
                            test.report.loc[_c_fail, "retry"] = "fail"
                    else:
                        test.report.loc[_c_fail, "retry"] = "boot_fail"
                else:
                    test.report.loc[_c_fail, "retry"] = "write_fail"
            except Exception:
                test.report.loc[_c_fail, "retry"] = "error"
        return core.Status.Pass
    except Exception:
        return core.Status.Fail
    finally:
        test.report_summary(os.path.join(SutConfig.Env.LOG_DIR, "unitool_variable_loop_summary.xlsx"))
        logging.info(f"Test finished after reboot {reboot_count} times")

