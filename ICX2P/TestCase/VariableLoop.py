import glob
import os
import re
import sys
import time
import json
import pandas
import random
import logging
import openpyxl
from pathlib import Path
from copy import deepcopy

from batf.SutInit import Sut
from batf.Report import ReportGen
from batf.Common.LogAnalyzer import LogAnalyzer
from batf import MiscLib, SshLib, SutInit, core, SerialLib, var
from ICX2P.Config import SutConfig
from ICX2P.Config.PlatConfig import Key, Msg, BootOS
from ICX2P.BaseLib import SetUpLib, PlatMisc, BmcLib, Update
from ICX2P.TestCase import UpdateBIOS


class BaseLine:
    def __init__(self, xlsx_file):
        self.xlsx = xlsx_file
        self.head_variable = "VariableName"
        self.head_value = "Value"
        self.head_default = "英文选择项"
        self.head_unitool = "UniTool"
        self.sheet_name = "X86 Setup Menu Baseline"
        self.supported_test_only = True
        self.unitool_base = 10  # 定义unitool读写的值是16进制还是10进制

    def load_sheet(self):
        work_book = openpyxl.load_workbook(self.xlsx)
        self.sheet = work_book.get_sheet_by_name(self.sheet_name)
        return self.sheet

    def get_header_data(self) -> dict:
        """
        return dict like:
        { value: column7, variable: column8, unitool: column13, default: column6}
        """
        self.load_sheet()
        header_index = {}
        for col in range(1, self.sheet.max_column + 1, 1):
            cell_value = self.sheet.cell(row=1, column=col).value
            if not cell_value:
                continue
            cell_value = cell_value.strip()
            if cell_value in [self.head_value, self.head_variable, self.head_unitool, self.head_default]:
                header_index[cell_value] = self.sheet.cell(row=1, column=col).col_idx
        return header_index

    def get_default_from_baseline(self):
        header = self.get_header_data()
        default_name = {}
        for irow in range(1, self.sheet.max_row + 1, 1):
            variable_name = self.sheet.cell(row=irow, column=header[self.head_variable]).value
            value_names = self.sheet.cell(row=irow, column=header[self.head_default]).value
            if not value_names or not variable_name:
                continue
            value_names, variable_name = value_names.strip(), variable_name.strip()
            for val in value_names.split("\n"):
                if not val.strip().endswith("*"):  # default value name end with character *
                    continue
                default_name[variable_name] = val.strip().strip("*").strip()
        return default_name

    def get_valid_data(self) -> dict:
        """return dict of all unitool support items like:
        {
        "variable1":
                { "value1": "0x0", "value2": "0x1" },
        "variable2":
                { "value1": "0x0", "value2": "0x1" }
        }
        """
        all_data = {}
        header_index = self.get_header_data()
        strikeout = 0
        non_stand = {}
        for irow in range(1, self.sheet.max_row + 1, 1):
            unicfg_support = self.sheet.cell(row=irow, column=header_index[self.head_unitool]).value
            variable_name = self.sheet.cell(row=irow, column=header_index[self.head_variable]).value
            value_int = self.sheet.cell(row=irow, column=header_index[self.head_value]).value
            value_default = self.sheet.cell(row=irow, column=header_index[self.head_default]).value

            unicfg_support = str(unicfg_support).strip() if unicfg_support else unicfg_support
            variable_name = str(variable_name).strip() if variable_name else variable_name
            value_int = str(value_int).strip() if value_int else value_int
            value_default = str(value_default).strip() if value_default else value_default

            if self.supported_test_only:
                if (not unicfg_support) or (not unicfg_support.strip().startswith("Y")):
                    continue
            else:
                if (not unicfg_support) or (not variable_name) or (not value_int) or (not value_default):
                    continue
            if all(self.sheet.cell(row=irow, column=co).font.strike for co in header_index.values()):
                logging.debug(f"Row {irow}:{variable_name} - is ignored due to strikeout")  # 忽略划了删除线的选项
                strikeout += 1
                continue
            value_int_lines = value_int.split("\n")
            val_spliter = [":", "："]  # 支持中英文冒号
            for val in value_int_lines:
                if any(spliter in val for spliter in val_spliter):
                    val_dict = []
                    for sp in val_spliter:
                        if sp not in val:
                            continue
                        val_k, val_v = val.split(sp, 1)
                        val_k = val_k.strip()
                        val_v = val_v.strip()
                        if "0x" in val_k.lower():
                            val_k, val_v = val_v, val_k
                        val_dict.append((val_k, val_v))
                    value_dict = dict(val_dict)
                    if variable_name in all_data:
                        all_data[variable_name].update(value_dict)
                    else:
                        all_data[variable_name] = value_dict
                else:
                    non_stand[variable_name] = val
        logging.debug(f"Total {strikeout} rows are strikeout")
        for non_k, non_v in non_stand.items():
            logging.warning(f"{non_k} = \"{non_v}\"")  # 不带冒号的Value选项
        return all_data

    def value_in_range(self) -> dict:
        """
        如过基线给定的Value类型是一个区间，就测试最大值和最小值
        return dict: 去掉 Max/Min 或 max/min 类选项的 step
        """
        self.valid_data = self.get_valid_data()
        idata = deepcopy(self.valid_data)
        for key, value in self.valid_data.items():
            for val in value.keys():
                if val.lower() == "step":
                    idata[key].pop(val)
        return idata

    def variable_in_range(self) -> dict:
        """
        如果基线给定的变量名是一个区间, 则将其实例化, 每个选项都单独列出来, BIOS基线整理后的原始数据字典
        list[0]~[3] 转为 list, list[1], list[2], list[3]
        list[0]~[3],[5] 转为 list, list[1], list[2], list[3], list[5]
        list[0]/[3]/[5] 转为 list, list[3]， list[5]
        return dict
        """
        data = self.value_in_range()
        idata = {}
        for key, value in data.items():
            if "[" not in key:
                idata[key] = value
                continue
            var_name = "".join(re.findall("(\w+)\[", key))
            var_index = re.findall("\[(\d+)]", key)
            if len(var_index) == 1 and ("[0]" in key):  # 去掉下表为[0]的索引
                idata[var_name] = value
            elif len(var_index) == 2 and ("~" in key):  # 正常的索引区间
                low, high = list(map(int, var_index))
                for i in range(low, high+1, 1):
                    if str(low) == "0":
                        idata[f"{var_name}"] = value
                        continue
                    idata[f"{var_name}[{i}]"] = value
            elif len(var_index) > 2:  # 非标准的索引区间
                range_index = re.findall("\[(\d+)]~\[(\d+)]", key)  # 区间部分
                if range_index:
                    low, high = list(map(int, range_index[0]))
                    for j in range(low, high+1, 1):
                        if str(low) == "0":
                            idata[f"{var_name}"] = value
                            continue
                        idata[f"{var_name}[{j}]"] = value
                other_index = re.findall("\[(\d+)]", re.sub("(\[\d+]~\[\d+])", "", key))  # 多个索引单独列出部分
                if other_index:
                    for o_index in other_index:
                        if str(o_index) == "0":
                            idata[f"{var_name}"] = value
                            continue
                        idata[f"{var_name}[{o_index}]"] = value
            else:  # 不带下标的索引
                idata[key] = value
        with open(os.path.join(SutConfig.Env.LOG_DIR, "origin_baseline.json"), "w", encoding="utf-8") as _json:
            json.dump(idata, _json, indent=4, ensure_ascii=False)
        return idata

    def make_test_data(self) -> dict:
        """
        生成字典 {变量名:[value1, value2...], 0x开头的转为十进制
        return dict like:
            {
                "variable1":["0x0", "0x1"],
                "variable2":["0x0", "0x1"],
            }
        """
        data = self.variable_in_range()
        idata = {}
        for var_name, values in data.items():
            for val_name, val_index in values.items():
                if val_index.lower().startswith("0x"):
                    val_index_hex = val_index
                elif not val_index.lower().isdigit():
                    logging.info(f"Baseline define value is not a number: [{var_name}:{val_index}]")
                    continue
                else:
                    logging.info(f"Baseline define as decimal, transfer to int: [{var_name}:{val_index}]")
                    # val_index_hex = int(hex(int(val_index, 10)))
                    val_index_hex = int(val_index, 10)
                if var_name in idata:
                    idata[var_name].append(val_index_hex)
                else:
                    idata[var_name] = [val_index_hex]
        return idata


class BaseLineLoop(BaseLine):
    def __init__(self, baseline, unitool, batch_count=20, default_json=None):
        super(BaseLineLoop, self).__init__(baseline)
        self.unitool = unitool
        self.batch_try_count = batch_count
        self.test_data = self.make_test_data()  # self.test_data的值格式为str: "0xff"
        self.init_report()
        self.read_default = self.read_default_value(local_json=default_json)  # self.read_default的值格式为str: "0xff"
        self.check_baseline_default()

    def init_report(self):
        report = {}
        for name, values in self.test_data.items():
            for val in values:
                report[f"{name}:{val}"] = {"default": None, "check_default": None, "error": None,
                                           "write": None, "check": None, "retry": None, "summary": None}
        self.report = pandas.DataFrame().from_dict(report, orient="index")

    def read_default_value(self, local_json=None, single_retry=False) -> dict:
        """
        通过unitool读取所有选项的当前默认value, 顺便检查变量名是否正确, 是否可读
        没有读到默认值的重试一次
        unitool读到的值默认为16进制
        """
        if local_json:
            with open(local_json, "r") as default_js:
                default = json.load(default_js, parse_int=str)
        else:
            default = self.unitool.read(*self.test_data)
            for df, val in default.items():
                if val is not None:
                    # hex_val = hex(int(val, self.unitool_base))
                    hex_val = str(int(val, self.unitool_base))
                    default.update({df: hex_val})
                    continue
                if single_retry:
                    retry_read = self.unitool.read(df)
                    if retry_read[df] is None:
                        logging.error(f"Read {df} error after retry")
                    else:
                        # hex_val_r = hex(int(retry_read[df], self.unitool_base))
                        hex_val_r = str(int(retry_read[df], self.unitool_base))
                        default.update({df: hex_val_r})
        for _df_key, _df_val in default.items():
            index_name = f"{_df_key}:{_df_val}"
            self.report.loc[index_name, "default"] = "yes"
            if _df_val is None:
                self.report.loc[index_name, "error"] = "yes"
        dump_read_file = os.path.join(SutConfig.Env.LOG_DIR, "unitool_read_default.json")
        with open(dump_read_file, "w") as read_json:
            json.dump(default, read_json, indent=4)
        return default

    def drop_default_value(self):
        """测试数据去掉通过 unitool 读取到的默认值, 如果无法读到默认值，说明变量异常，跳过测试"""
        for name, default in self.read_default.items():
            try:
                self.test_data[name].remove(default)
            except Exception:
                logging.warning(f"Default not in values list: {name} = {default}")

    def check_baseline_default(self):
        """检查基线的默认值和uniCfg读到的默认值是否一致，需要先用uniCfg将所有的变量的值读出来再做比较"""
        unicfg_read_default = self.read_default
        base_default = self.get_default_from_baseline()
        for op_name, val_name in base_default.items():
            read_int_str = unicfg_read_default.get(op_name)
            if not read_int_str:
                continue
            if not self.valid_data.get(op_name):
                logging.info(f"Baseline default attribute name missing: {op_name}")
                continue
            base_value_str = self.valid_data.get(op_name).get(val_name)
            if not base_value_str:
                logging.info(f"Baseline default attribute name: {op_name} missing value name: {base_value_str}")
                continue
            base_int = int(base_value_str, 16) if base_value_str.startswith("0x") else int(base_value_str)
            index_name = f"{op_name}:0x{base_int}" if base_value_str.startswith("0x") else f"{op_name}:{base_int}"
            if int(read_int_str, self.unitool_base) != base_int:
                logging.info(f"Check baseline default fail, {op_name}: baseline={base_value_str} read={read_int_str}")
                self.report.loc[index_name, "check_default"] = "fail"
                continue
            self.report.loc[index_name, "check_default"] = "pass"

    def testdata_generator(self, test_scope=None):
        """生成器，每次调用时，临时生成数量为 {self.batch_try_count} 的测试数据字典，用于批量测试"""
        test_scope = [] if not test_scope else test_scope
        check_data = deepcopy(self.test_data)
        while check_data:
            test_count = 0
            temp_data = {}
            for name in self.test_data:
                if name not in check_data:  # 变量全部测完，跳过
                    continue
                if self.read_default.get(name) is None:  # 读取变量名报错的跳过
                    check_data.pop(name)
                    continue
                if not check_data[name]:  # value全部测完，删除整个变量
                    check_data.pop(name)
                    continue
                value_choice = random.choice(check_data[name])
                if test_scope and (f"{name}:{value_choice}" not in test_scope):  # 已经有测试结果的跳过
                    check_data[name].remove(value_choice)
                    continue
                temp_data[name] = value_choice
                check_data[name].remove(value_choice)  # 单个value测完后删除此value
                test_count += 1
                if test_count >= self.batch_try_count:
                    break
            if temp_data:
                yield temp_data

    def single_write(self, data):
        """输入一个或多个选项，一项一项单独修改,并记录write结果"""
        for sw_key, sw_val in data.items():
            index_name = f"{sw_key}:{sw_val}"
            try:
                write_reslut = self.unitool.write(**{sw_key: sw_val})
                if write_reslut:
                    self.report.loc[index_name, "write"] = "pass"
                else:
                    self.report.loc[index_name, "write"] = "fail"
            except Exception:
                self.report.loc[index_name, "write"] = "error"

    def batch_write(self, data: dict):
        """输入一个或多个选项，批量修改，修改失败或报错再尝试单个修改, 并记录write结果"""
        try:
            if self.unitool.write(**data):
                for bw_key, bw_val in data.items():
                    index_name = f"{bw_key}:{bw_val}"
                    self.report.loc[index_name, "write"] = "pass"
            else:
                logging.info("batch write fail, retry write one by one...")
                self.single_write(data)
        except Exception:
            logging.info("batch write error, retry write one by one...")
            self.single_write(data)

    def single_check(self, data):
        """输入一个或多个选项，一项一项单独检查,并记录check结果"""
        for sc_key, sc_val in data.items():
            index_name = f"{sc_key}:{sc_val}"
            try:
                check_reslut = self.unitool.check(**{sc_key: sc_val})
                assert (check_reslut is not None)
                if check_reslut:
                    self.report.loc[index_name, "check"] = "pass"
                else:
                    self.report.loc[index_name, "check"] = "fail"
            except Exception:
                self.report.loc[index_name, "check"] = "error"

    def batch_check(self, data: dict):
        """输入一个或多个选项，批量检查，检查失败或报错再尝试单个检查, 并记录check结果"""
        try:
            if self.unitool.check(**data):
                for bc_key, bc_value in data.items():
                    index_name = f"{bc_key}:{bc_value}"
                    self.report.loc[index_name, "check"] = "pass"
            else:
                logging.info("batch check fail, retry check one by one...")
                self.single_check(data)
        except Exception:
            logging.info("batch check error, retry check one by one...")
            self.single_check(data)

    def report_summary(self, excel_save: str):
        self.report.sort_index(inplace=True)
        self.report["summary"] = self.report.apply(lambda x:
            "pass" if ((x.write == "pass" and x.check == "pass") or x.retry == "pass") else
            ("fail" if (x.write == "fail" or x.check == "fail" or x.retry == "fail") else None), axis=1)
        if excel_save.endswith(".xlsx"):
            self.report.to_excel(excel_save)

    def items_not_tested(self, xlsx=None):  # 测到一半时若中断，将测试上次的报告复制到SetupBase文件夹继续测试
        if xlsx:
            self.report = pandas.read_excel(xlsx, index_col=0)
        test_list = []
        pd_test_index = self.report.index.to_list()
        for index in pd_test_index:
            if pandas.isna(self.report.loc[index, "write"]) and pandas.isna(self.report.loc[index, "default"]):
                test_list.append(index)
        return test_list

    def get_test_counts(self, test_scope=None):
        n = 0
        for _ in self.testdata_generator(test_scope):
            n += 1
        return n


@core.test_case(("123", "[TC123] Uitool Variables Loop Test", "Uitool Variables Loop Test"))
def unicfg_variable_loop_test():
    baseline = PlatMisc.root_path() / "ICX2P/Tools/SetupBase/2288服务器setup菜单基线.xlsx"
    reboot_count = 0

    def reboot_ping(reboot=True, timeout=720):
        nonlocal reboot_count
        if reboot:
            BmcLib.force_reset()
            time.sleep(30)
            reboot_count += 1
        if MiscLib.ping_sut(SutConfig.Env.OS_IP, timeout):
            return True
        BmcLib.force_reset()  # 避免开机卡在 PostCode: 63
        reboot_count += 1
        return MiscLib.ping_sut(SutConfig.Env.OS_IP, timeout)

    def flash_bios():  # 刷BIOS并开机
        logging.info("Unable to boot up after setting, flash bios image to restore")
        if not var.get("biosimage"):
            img = Update.get_test_image(SutConfig.Env.LOG_DIR, "master", 'debug-build')
            var.set("biosimage", img)
        UpdateBIOS.update_bios(var.get("biosimage"))
        BmcLib.force_reset()

    try:
        # BmcLib.clear_cmos()
        # assert reboot_ping(), "Boot to os failed"

        # 若已经有Unitool读取到的默认值json，就不再通过unitool重复读取，节省时间
        # local_default_json = PlatMisc.root_path()/"Resource/SetupBase/unitool_read_default.json"
        local_default_json = Path(r"D:\2001\2288V6\automation\ICX2P\Tools\SetupBase\unitool_read_default.json")
        default_json = local_default_json if local_default_json.exists() else None
        test = BaseLineLoop(baseline, unitool=Sut.UNITOOL, batch_count=30, default_json=default_json)
        test.drop_default_value()  # 默认值直接跳过不测

        # 若已经有一部分测试结果，依据之前的测试结果继续测试，节省时间
        items_tested_xlsx = Path(r"D:\2001\2288V6\automation\ICX2P\Tools\SetupBase\unitool_variable_loop_summary.xlsx")
        items_not_tested = test.items_not_tested(items_tested_xlsx) if items_tested_xlsx.exists() else None

        # 开始批量测试
        all_test_count = test.get_test_counts(test_scope=items_not_tested)
        logging.info(f"Total need loop test {all_test_count} times")
        batch_process = 1
        for data in test.testdata_generator(test_scope=items_not_tested):  # 生成批量测试数据
            logging.info(f"{'='*50}[{batch_process}/{all_test_count}]")  # 展示批量测试进度
            if reboot_ping(reboot=False):
                test.batch_write(data)
                write_pass = {}  # 只有write成功的才去检查
                for data_k, data_v in data.items():
                    if test.report.loc[f"{data_k}:{data_v}", "write"] == "pass":
                        write_pass[data_k] = data_v
                if not write_pass:  # 如果批量全部都设置失败，则开始下一组测试
                    continue
                if reboot_ping():
                    test.batch_check(write_pass)
                else:
                    for one_key, one_value in write_pass.items():  # write pass但重启无法开机的标为 boot_fail
                        test.report.loc[f"{one_key}:{one_value}", "check"] = "boot_fail"
            BmcLib.clear_cmos()  # 每组测试完成后clear CMOS
            if not reboot_ping():  # 如果修改后ClearCMOS也无法开机，则刷BIOS恢复
                flash_bios()
            batch_process += 1
        test.report_summary(os.path.join(SutConfig.Env.LOG_DIR, "unitool_batch_test.xlsx"))

        # 批量测试失败的单独测试
        check_fail = test.report[test.report["check"].str.contains("fail", na=False)]
        fail_count = len(check_fail.index)
        fail_process = 1
        is_writed = True
        for fail_name in check_fail.index:
            logging.info(f"{'='*50}[{fail_process}/{fail_count}]")
            logging.info(f"Start to single retry {fail_name}")  # 显示测试进度
            fail_process += 1
            if is_writed:  # 如果unitool没有write成功，就不需要clearCMOS并重启
                BmcLib.clear_cmos()
                if not reboot_ping():  # 如果修改后ClearCMOS也无法开机，则刷BIOS恢复
                    flash_bios()
                is_writed = False
            if reboot_ping(reboot=False):
                try:
                    item_retry = dict([fail_name.rsplit(":", 1)])
                    if not Sut.UNITOOL.write(**item_retry):
                        test.report.loc[fail_name, "retry"] = "write_fail"
                        continue
                    is_writed = True
                    if not reboot_ping():
                        test.report.loc[fail_name, "retry"] = "reboot_fail"
                        continue
                    if not Sut.UNITOOL.check(**item_retry):
                        test.report.loc[fail_name, "retry"] = "fail"
                        continue
                    test.report.loc[fail_name, "retry"] = "pass"
                except Exception as e:
                    test.report.loc[fail_name, "retry"] = f"error:{e}"
        return core.Status.Pass
    except Exception as e:
        logging.error(e)
        return core.Status.Fail
    finally:
        test.report_summary(os.path.join(SutConfig.Env.LOG_DIR, "unitool_variable_loop_summary.xlsx"))
        logging.info(f"Test finished after reboot {reboot_count} times")


@core.test_case(("124", "[TC124] Uitool Check Baseline default value", "Uitool Check Baseline default value"))
def baseline_default_check():
    baseline = PlatMisc.root_path() / "ICX2P/Tools/SetupBase/2288服务器setup菜单基线.xlsx"
    try:
        BmcLib.clear_cmos()
        assert BmcLib.force_reset()
        assert MiscLib.ping_sut(SutConfig.Env.OS_IP, 600)
        test = BaseLineLoop(baseline, unitool=Sut.UNITOOL, batch_count=30, default_json=None)
        return core.Status.Pass
    except Exception as e:
        logging.error(e)
        return core.Status.Fail
    finally:
        test.report_summary(os.path.join(SutConfig.Env.LOG_DIR, "baseline_default_value_test.xlsx"))

