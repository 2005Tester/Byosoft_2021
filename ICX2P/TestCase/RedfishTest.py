import os
import re
import time
import json
import random
import pandas
import logging
import openpyxl
from copy import deepcopy
from pathlib import Path

from batf.SutInit import Sut
from batf.Common.RedfishLib import Redfish
from batf.Report import ReportGen
from batf.Common.LogAnalyzer import LogAnalyzer
from batf import MiscLib, SshLib, SutInit, core, SerialLib
from ICX2P.Config import SutConfig
from ICX2P.Config.PlatConfig import Key, Msg, BootOS
from ICX2P.BaseLib import SetUpLib, PlatMisc, BmcLib

base_line = Path(__file__).parent.parent.absolute()/"Tools/SetupBase/2288服务器setup菜单基线版本_Byosoft_V0.1.xlsx"
power_efficiency = Path(__file__).parent.parent.absolute()/"Tools/PowerEfficiency/2288V6_PowerEfficiency.csv"


class _BootInvolved:
    Exclusive = ["BootTypeOrder0", "BootTypeOrder1", "BootTypeOrder2", "BootTypeOrder3"]
    MemPop = ["MemChannelEnable"]
    BootOS = ["BootType", "PowerOnPassword"]
    ReadOnly = ["OemSecureBoot"]


def _to_excel(data, name="", path=SutConfig.Env.LOG_DIR):
    """列表，字典或DataFrame转成excel文件"""
    if not os.path.exists(path):
        os.makedirs(path)
    now = time.strftime("%Y%m%d_%H%M%S", time.localtime())
    path_name = os.path.join(path, rf"{name} {now}.xlsx")
    if isinstance(data, list):
        pandas.Series(data).to_excel(path_name)
    elif isinstance(data, dict):
        pandas.DataFrame.from_dict(data, orient='index').to_excel(path_name)
    elif isinstance(data, pandas.DataFrame):
        data.to_excel(path_name)
    logging.info("Create excel {} successful!".format(path_name))


def _read_default_values(variables):
    """通过Unitool读取多个变量的值，返回字典"""
    var_format = []
    for var in variables:
        if "[0]" in var:
            var_format.append(var[:var.find("[")])
        else:
            var_format.append(var)
    key_values = Sut.UNITOOL.read(*var_format)
    return key_values


def _reboot_sut(timeout=600):
    """重启到BIOS结束，不需要进OS"""
    BmcLib.force_reset()
    if SerialLib.is_msg_present(Sut.BIOS_COM, msg=Msg.BIOS_BOOT_COMPLETE, delay=timeout):
        SerialLib.clean_buffer(Sut.BIOS_COM)
        return True


def _option_value_index_map(base_xlsx):
    """从基线中解析出{变量名:值}的字典，返回字典"""
    base_xlsx = openpyxl.load_workbook(base_xlsx)
    sheets = base_xlsx.sheetnames
    base_sheet = base_xlsx[sheets[0]]
    setup_col = "K"  # setup column index of xlsx
    value_col = "I"  # value column index of xlsx
    max_row = base_sheet.max_row
    table = {}
    for row in range(1, max_row):
        option = base_sheet[f"{setup_col}{row}"].value
        if not option:
            continue
        option = option.replace("\n", "").strip()
        values = base_sheet[f"{value_col}{row}"].value
        if not values:
            continue
        value_list = values.split("\n")
        for v in value_list:
            if len(re.findall(":", v)) == 1:
                vname = v.split(":")[0]
                vidx = v.split(":")[1].strip()
                index = str(int(vidx, 16)) if "0x" in vidx else vidx
                if isinstance(table.get(option), dict):
                    table[option].update({index: vname.strip()})
                else:
                    table[option] = {index: vname.strip()}
            elif len(re.findall(":", v)) > 1:
                vname = v[:v.rfind(":")]
                vidx = v[v.rfind(":")+1:]
                index = str(int(vidx, 16)) if "0x" in vidx else vidx
                if isinstance(table.get(option), dict) and table.get(option):
                    table[option].update({index: vname.strip()})
                else:
                    table[option] = {index: vname.strip()}
            else:
                table[option] = v
    return table


class _DepenTest(Redfish):
    """
    【测试配置】
    1. 配置 config 文件
        -   bmc_ip
        -   bmc_user
        -   bmc_pw
        -   COM
    2. 进入到Redfish文件夹，运行命令 RedFishTest.py depmain 开始测试
    3. 测试完成后Excel报告保存在 config.REPORT_DIR 路径中

    【测试流程】
    1. Dump Registry文件
    2. 以MapFrom列表为索引，整理出每个MapFrom条件，关联的下级MapTo菜单，作为一行
    3. 去掉所有的反向关联条件（Hidden=False)
    4. 遍历DataFrame的每一行，并尝试PATCH MapFrom 列表
        a.  每次遍历前先通过POST恢复默认
        b.  每次遍历都去检查此菜单是否有多级联动导致不能直接修改的情况，如果有，则让此条件和上级菜单（选择一个可修改的值）一起PATCH
    5. 修改完成后重启
    6. 重启后检查 MapFrom 的值有没有修改成功
    7. 然后检查每一个MapTo的条件，查看是否满足
        a.  CurrentValue的情况：    Value为PATCH的值则认为 pass
        b.  Hidden/ReadOnly=True的情况:    尝试PATCH并PATCH失败则认为 pass
        c.  其他情况:   尝试PATCH并PATCH成功则认为 pass
    8.  生成Excel报告
        PASS:   -  MapFrom PATCH成功
                -  重启后检查MapFrom的值符合要求
                -  重启后检查MapTo的值或者属性符合要求
        FAIL:   -  任意一项为Fail
    """

    def __init__(self):
        super(_DepenTest, self).__init__(SutConfig.Env.BMC_IP, SutConfig.Env.BMC_USER, SutConfig.Env.BMC_PASSWORD, SutConfig.Env.REDFISH_API)
        self.atts_pd = pandas.DataFrame()
        self.map_from_pd = pandas.DataFrame()
        self.depens = self.Dependencies()
        self.mapto_link = self.mapto_info()
        self.registry_dump(dump_json=True, path=SutConfig.Env.LOG_DIR)

    def attributes_info(self, dump: bool = False):
        """ 以AttributeName为索引返回 DataFrame，可选择生成Excel文件用于手动检查 """
        atts = self.Attributes()
        for index, att in enumerate(atts):
            if atts[index].get("Value"):
                atts[index]["Value"] = [val["ValueName"] for val in att["Value"]]
        self.atts_pd = pandas.DataFrame(atts).set_index("AttributeName")
        if dump:
            _to_excel(self.atts_pd, "AttributesName_所有信息")
        return self.atts_pd

    def depenfor_info(self, dump: bool = False):
        """ 以DependencyFor为索引返回 DataFrame，可选择生成Excel文件用于手动检查 """
        depen_pd = pandas.DataFrame(self.depens).set_index("DependencyFor")
        depen_pd["Dependency"] = depen_pd["Dependency"].apply(lambda x: json.dumps(x, indent=4))
        if dump:
            _to_excel(depen_pd, "DependencyFor_检索信息")
        return depen_pd

    def mapto_info(self, dump: bool = False):
        """ 以MapToAttribute为索引返回 DataFrame，可选择生成Excel文件用于手动检查 """
        depen_dic = {}
        depen_link = {}  # used for map link in self.mapped_as_unchanged()
        for depen in self.depens:
            map_to_name = depen["Dependency"]["MapToAttribute"]
            for d in depen["Dependency"]["MapFrom"]:
                mt_property = depen["Dependency"]["MapToProperty"]
                mt_value = depen["Dependency"]["MapToValue"]
                mf_name = d["MapFromAttribute"]
                mf_value = d["MapFromValue"]
                if not depen_dic.get(map_to_name):
                    depen_dic[map_to_name] = [depen]
                else:
                    depen_dic[map_to_name].append(depen)
                if (mt_property == "Hidden" or mt_property == "ReadOnly") and mt_value:
                    if not depen_link.get(map_to_name):
                        depen_link.setdefault(map_to_name, {})
                    if not depen_link[map_to_name].get(mf_name):
                        depen_link[map_to_name].setdefault(mf_name, [])
                    depen_link[map_to_name][mf_name].append(mf_value)
        map_to_pd = pandas.DataFrame.from_dict(depen_dic, orient="index").applymap(lambda x: json.dumps(x, indent=4))
        if dump:
            _to_excel(map_to_pd, "MapToAttribute_检索信息")
        return depen_link

    def mapfrom_info(self, dump: bool = False):
        """ 以MapFrom为索引返回 DataFrame，可选择生成Excel文件用于手动检查 """
        map_from_dic = {}
        deps = deepcopy(self.depens)
        for depen in deps:
            mapf = str(depen["Dependency"]["MapFrom"])  # "[{},{}]"
            depen["Dependency"].pop("MapFrom")
            mapt = depen["Dependency"]  # {}
            if not (mapf in map_from_dic):
                map_from_dic[mapf] = [mapt]  # [{}]
            else:
                map_from_dic[mapf].append(mapt)  # # [{},{}]
        self.map_from_pd = pandas.DataFrame.from_dict(map_from_dic, orient="index").reset_index()
        mf_depen_relation = self.map_from_pd
        mf_depen_relation["index"] = mf_depen_relation["index"].apply(eval)
        mf_depen_relation = mf_depen_relation.applymap(lambda x: json.dumps(x, indent=4))
        if dump:
            _to_excel(mf_depen_relation, "MapFrom_MapTo_联动关系")
        return self.map_from_pd

    def drop_hidden_false(self, dump: bool = False):
        """ 去除所有 Hidden=false的反向关系 """
        hid_false = re.compile(r"MapToProperty.+Hidden.+MapToValue.+False")
        for column_index, column_data in self.map_from_pd.iteritems():
            for row_index, row_data in column_data.iteritems():
                if hid_false.search(str(row_data)) or (str(row_data) == "None"):
                    self.map_from_pd.loc[row_index, column_index] = None
        self.map_from_pd.dropna(subset=[n for n in range(0, len(self.map_from_pd.columns) - 1)], how="all",
                                inplace=True)
        # self.map_from_pd["index"] = self.map_from_pd["index"].apply(eval)
        self.map_from_pd.reset_index(inplace=True, drop=True)
        mf_hidden_false = self.map_from_pd.applymap(lambda x: json.dumps(x, indent=4))
        if dump:
            _to_excel(mf_hidden_false, "MapFrom_去除反向联动信息")
        return self.map_from_pd

    def mapped_as_unchanged(self, mapfrom: list):
        """ 检查选项是否会被关联并且不允许修改 (Hidden=True/ReadOnly=True) """
        result_mapto = True
        for index, mf in enumerate(mapfrom):
            mf_name = mapfrom[index]["MapFromAttribute"]
            for mk in self.mapto_link.keys():
                if mf_name == mk:
                    result_mapto = result_mapto & False
        return not result_mapto

    @staticmethod
    def gen_patch_data(map_from: list):
        """
        生成PATCH的字典
            1. Only one item just return       {key1:value1}
            2. One more items
                MapTerms = AND return all       {key1:value1, key2:value2, ...}
                MapTerms = OR return 1st one    {key1:value1}
        """
        option = {}
        for index, mf in enumerate(map_from):
            mf_key = map_from[index]["MapFromAttribute"]
            mf_value = map_from[index]["MapFromValue"]
            option[mf_key] = mf_value
            if len(map_from) == 1:
                return option
            if map_from[index]["MapTerms"] == "OR":
                return option
        return option

    def not_the_value(self, mf_key, mf_value, mt_key):
        vals = deepcopy(self.atts_pd.loc[mf_key, "Value"])
        mt_link = deepcopy(self.mapto_link[mt_key][mf_key])
        if mf_value in vals:
            vals.remove(mf_value)  # 去掉MapFrom的值
        if len(vals) == 1:
            return random.choice(vals)
        for mk in mt_link:
            if mk in vals:
                vals.remove(mk)  # 去掉被关联且不能改的值
        return random.choice(vals)

    @staticmethod
    def mf_key_value(mapfrom: list):
        """ 根据MapFrom生成字典 {MapFromAttribute：MapFromValue} """
        mf_dic = {}
        for mf in mapfrom:
            mf_dic.update({mf["MapFromAttribute"]: mf["MapFromValue"]})
        return mf_dic

    def gen_data_w_top(self, mapfrom: list):
        """ 当发现 MapFromAttribute 被联动时，找到上级联动一起修改 """
        mf_dic = self.mf_key_value(mapfrom)
        depens = deepcopy(self.depens)
        for depen in depens:
            mapto_name = depen["Dependency"]["MapToAttribute"]
            if mapto_name not in mf_dic.keys():
                continue
            mapto_property = depen["Dependency"]["MapToProperty"]
            mapto_value = depen["Dependency"]["MapToValue"]
            if not ((mapto_property == "Hidden" or mapto_property == "ReadOnly") and mapto_value):
                continue
            for mf in depen["Dependency"]["MapFrom"]:
                key = mf["MapFromAttribute"]
                if key in mf_dic.keys():  # 避免PATHC Value被覆盖修改
                    continue
                value = self.not_the_value(key, mapto_value, mapto_name)
                kv = {key: value}
                if mf["MapTerms"] == "OR":  # OR的只选择一个
                    mf_dic.update(kv)
                    break
                mf_dic.update(kv)  # AND的选择所有
        return mf_dic

    def check_mapto(self, mapfrom: list):
        """ PATCH完成后重启，检查所有的MapTo的联动条件是否满足 """
        # pandas DataFrame filter doesn't support list types, use column IndexStr instead
        mf_info = {mf["MapFromAttribute"]: mf["MapFromValue"] for mf in mapfrom}
        mapto_row = self.map_from_pd[self.map_from_pd["IndexStr"].values == json.dumps(mapfrom, indent=4)].iloc[0]
        result = True
        for mi, mt in mapto_row.items():
            if not isinstance(mi, int):
                continue
            if not mt:
                continue
            map_key = mt.get("MapToAttribute")
            map_property = mt.get("MapToProperty")
            map_value = mt.get("MapToValue")
            # json transfer data as str, redfish may get value as int
            if isinstance(map_value, str):
                map_value = int(map_value) if map_value.isdigit() else map_value
            # Check Value Equal MapTo
            if map_property == "CurrentValue":
                if not self.check_bios_option(**{map_key: map_value}):
                    result = result & False
                    logging.info(rf"[CHECK] MapTO: {map_key}: {map_property} -> {map_value} <fail>")
                    continue
                logging.info(rf"[CHECK] MapTO: {map_key}: {map_property} -> {map_value} <pass>")
                continue
            # Hidden/ReadOnly map_key can't be changed
            patch_try = self.patch_radom(map_key)
            if patch_try.result is None:
                continue
            if (map_property == "Hidden" or map_property == "ReadOnly") and map_value:
                if patch_try.result:
                    result = result & False  # Hidden/ReadOnly should patch fail
                    logging.info(rf"[CHECK] MapTO: {map_key}: {map_property} -> {map_value} <fail>")
                    logging.info(rf"Status: {patch_try.status}")
                    continue
                logging.info(rf"[CHECK] MapTO: {map_key}: {map_property} -> {map_value} <pass>")
                continue
            # is map_key be mapped bt other item as Hidden/ReadOnly ?
            if self.mapto_link.get(map_key):
                for kmf, vmf in mf_info.items():
                    if vmf in self.mapto_link[map_key].get(kmf):
                        if patch_try.result:
                            result = result & False
                            logging.info(rf"[CHECK] MapTO: {map_key}: {map_property} -> {map_value} <fail>")
                            logging.info(rf"Message: {patch_try.body}")
                            continue
                        logging.info(rf"[CHECK] MapTO: {map_key}: {map_property} -> {map_value} <pass>")
                    else:
                        if not patch_try.result:
                            result = result & False
                            logging.info(rf"[CHECK] MapTO: {map_key}: {map_property} -> {map_value} <fail>")
                            logging.info(rf"Message: {patch_try.body}")
                            continue
                        logging.info(rf"[CHECK] MapTO: {map_key}: {map_property} -> {map_value} <pass>")
                continue
            # Others Patch pass
            if not patch_try.result:
                result = result & False
                logging.info(rf"[CHECK] MapTO: {map_key}: {map_property} -> {map_value} <fail>")
                logging.info(rf"Message: {patch_try.body}")
                continue
            logging.info(rf"[CHECK] MapTO: {map_key}: {map_property} -> {map_value} <pass>")
        return result

    def patch_radom(self, key: str):
        """ 随机选择一个有效Value去PATCH，用于重启后检查是否能修改，判断是否隐藏或只读 """
        attpd = self.atts_pd
        if key not in attpd.index:
            logging.info(rf"[{key}] not in AttributeName list, <skipped>")
            return {"result": None}
        att_series = attpd.loc[key]
        default = att_series["DefaultValue"]
        if att_series["Type"] == "Enumeration":
            vale = deepcopy(att_series["Value"])
            if default in vale:
                vale.remove(default)
            value = random.choice(vale)
            return self.set_bios_option(**{key: value})
        if att_series["Type"] == "Integer":
            low = int(att_series["LowerBound"])
            up = int(att_series["UpperBound"] + 1)
            scrlar = int(att_series["ScalarIncrement"])
            scrlar = 1 if (scrlar == 0) else scrlar
            iter_range = list(range(low, up, scrlar))
            iter_range.remove(default)
            ival = random.choice(iter_range)
            return self.set_bios_option(**{key: ival})

    def gen_report(self):
        """ 生成最终结果 """
        self.map_from_pd["Summary"] = self.map_from_pd.apply(
            lambda x:
            "Pass" if (
                              x.CheckStatus == "pass" and x.PatchStatus == "pass" and x.CheckMapTo == "pass")  else
            ("Fail" if (
                    x.CheckStatus == "fail" or x.PatchStatus == "fail" or x.CheckMapTo == "fail") else
             None), axis=1)
        self.map_from_pd.drop(labels="IndexStr", axis=1, inplace=True)
        self.map_from_pd["index"] = self.map_from_pd["index"].apply(json.dumps).apply(eval)
        self.map_from_pd["index"] = self.map_from_pd["index"].apply(lambda x: json.dumps(x, indent=4))
        _to_excel(self.map_from_pd, "DepMainSummary")

    def run_test(self):
        # 初始化整理数据
        self.depenfor_info(dump=True)
        self.mapto_info(dump=True)
        self.attributes_info(dump=True)
        self.mapfrom_info(dump=True)
        self.drop_hidden_false(dump=True)
        pdlist = self.map_from_pd.columns.tolist()
        pdlist = pdlist[:1] + ["IndexStr", "PATCH", "IsMapped", "PatchStatus", "CheckStatus", "CheckMapTo",
                               "Summary"] + pdlist[1:]
        self.map_from_pd = self.map_from_pd.reindex(columns=pdlist)
        self.map_from_pd["IndexStr"] = self.map_from_pd["index"].apply(lambda x: json.dumps(x, indent=4))

        # 遍历测试
        test_count = 0
        for row, mf_item in self.map_from_pd["index"].iteritems():
            test_count += 1
            try:
                logging.info(f"==========================================================[{test_count}/{len(self.map_from_pd['index'])}]")

                # Load Default and Reboot
                self.bios_load_default()
                if not _reboot_sut():
                    logging.info("Boot up failed")
                    continue
                logging.info("Boot up successfully")

                # print test title
                key_value = self.gen_patch_data(mf_item)
                logging.info("Start test {}".format(key_value))

                # PATCH items of mapped as unchanged
                if self.mapped_as_unchanged(mf_item):
                    logging.info(r"Items is mapped as Hidden or ReadOnly, try patch with top-level")
                    # 多级联动的情况，将主菜单和子菜单一起修改，主菜单要改为 非 Hidden / ReadOnly
                    self.map_from_pd.loc[row, "IsMapped"] = "yes"
                    key_value = self.gen_data_w_top(mf_item)
                    self.map_from_pd.loc[row, "PATCH"] = json.dumps(key_value, indent=4)
                    patch_result = self.set_bios_option(**key_value)
                    if not patch_result.result:
                        logging.info(r"[PATCH] {} <fail>".format(key_value))
                        logging.info(f"Status: {patch_result.status}")
                        logging.info(rf"Message: {patch_result.body}")
                        self.map_from_pd.loc[row, "PatchStatus"] = "fail"
                        continue
                    logging.info(r"[PATCH] {} <pass>".format(key_value))
                    logging.info(rf"Status: {patch_result.status}")
                    self.map_from_pd.loc[row, "PatchStatus"] = "pass"
                else:
                    # Normal PATCH
                    self.map_from_pd.loc[row, "PATCH"] = json.dumps(key_value, indent=4)
                    patch_result = self.set_bios_option(**key_value)
                    if not patch_result.result:
                        logging.info(r"[PATCH] {} <fail>".format(key_value))
                        logging.info(rf"Error Message: {patch_result.body}")
                        self.map_from_pd.loc[row, "PatchStatus"] = "fail"
                        continue
                    logging.info(r"[PATCH] {} <pass>".format(key_value))
                    self.map_from_pd.loc[row, "PatchStatus"] = "pass"

                # REBOOT
                if not _reboot_sut():
                    logging.info("Boot up failed")
                    continue
                logging.info("Boot up successfully")

                # CHECK MapFrom
                if not self.check_bios_option(**key_value):
                    logging.info(r"[CHECK] >>> MapFrom: {} <fail>".format(key_value))
                    self.map_from_pd.loc[row, "CheckStatus"] = "fail"
                    continue
                logging.info(r"[CHECK] >>> MapFrom: {} <pass>".format(key_value))
                self.map_from_pd.loc[row, "CheckStatus"] = "pass"

                # CHECK MapTo
                if not self.check_mapto(mf_item):
                    self.map_from_pd.loc[row, "CheckMapTo"] = "fail"
                    logging.info(r"[CHECK] >>> MapTo Result: <fail>")
                    continue
                self.map_from_pd.loc[row, "CheckMapTo"] = "pass"
                logging.info(r"[CHECK] >>> MapTo Result: <pass>")

                # Force to UEFI if 'BootType' is set to 'LegacyBoot' (this attribute can't be load default with post)
                if ('BootType', 'LegacyBoot') in key_value.items():
                    assert self.set_bios_option(BootType='UEFIBoot').result
                    assert _reboot_sut()

            except Exception as e:
                logging.info(e)
                continue

        # Result Summary
        self.gen_report()


class _NonDepTest(Redfish):
    def __init__(self):
        super(_NonDepTest, self).__init__(SutConfig.Env.BMC_IP, SutConfig.Env.BMC_USER, SutConfig.Env.BMC_PASSWORD, SutConfig.Env.REDFISH_API)
        self.registry_dump(dump_json=True, path=SutConfig.Env.LOG_DIR)

    def attributes_data(self):
        atts = self.Attributes()
        for index, att in enumerate(atts):
            if atts[index].get("Value"):
                atts[index]["Value"] = [val["ValueName"] for val in att["Value"]]
        self.att_pd = pandas.DataFrame(atts).set_index("AttributeName")
        logging.info(f"Current Registry.json contains {len(self.att_pd.index)} attributes names")

    def get_non_default_value(self, key):
        default_value = self.att_pd.loc[key, "DefaultValue"]
        if default_value is None:  # 没有默认值的情况
            logging.info(f"[Warning] DefaultValue: {key} = {default_value}")
            return
        if self.att_pd.loc[key, "Type"] == "Enumeration":  # 枚举型值
            values = deepcopy(self.att_pd.loc[key, "Value"])
            if len(values) < 2:
                logging.info(f"[Warning] Only One Value: {key} = {values}")
                return
            values.remove(default_value)
            return random.choice(values)
        if self.att_pd.loc[key, "Type"] == "Integer":  # 整数型值
            scalar = deepcopy(self.att_pd.loc[key, "ScalarIncrement"])
            scalar = scalar if scalar else 1
            lower = deepcopy(self.att_pd.loc[key, "LowerBound"])
            upper = deepcopy(self.att_pd.loc[key, "UpperBound"])
            values = list(range(int(lower), int(upper) + 1, int(scalar)))
            values.remove(default_value)
            return int(random.choice(values))

    def gen_patch_key_value(self):
        self.patch_key_value = {}
        for att in self.att_pd.index:
            value = self.get_non_default_value(att)
            if value is not None:
                self.patch_key_value[att] = value
        self.patch_key_value.update(self.boot_order_key_value())  # Boot Order取值需要特殊处理

    def report_init(self):
        self.result = pandas.DataFrame.from_dict(self.patch_key_value, orient="index")  # 整理报告格式
        self.result.rename(columns={0: "patch_value"}, inplace=True)
        pdlist = self.result.columns.tolist()
        col_list = ["default_value", "patch_status", "check_status", "summary", "message"]
        col_list = col_list[:1] + pdlist + col_list[1:]
        self.result = self.result.reindex(columns=col_list)

    def drop_been_mapped(self):
        """ CurrentValue / Hidden=True / ReadOnly=True """
        dep_data = self.Dependencies()
        self.been_mapto = {}
        for dep in dep_data:
            mapfrom = {mf["MapFromAttribute"]: mf["MapFromValue"] for mf in dep["Dependency"]["MapFrom"]}
            for mf_key, mf_value in mapfrom.items():
                try:
                    if self.att_pd.loc[mf_key, "DefaultValue"] != mf_value:  # 有联动关系但默认情况不关联的视为非关联项，可直接修改
                        continue
                except Exception as e:
                    logging.error(f"{e} missing in attributes list")
            mapto_name = dep["Dependency"]["MapToAttribute"]
            mapto_property = dep["Dependency"]["MapToProperty"]
            mapto_value = dep["Dependency"]["MapToValue"]
            if (mapto_property == "CurrentValue") or ((mapto_property in ["Hidden", "ReadOnly"]) and mapto_value):
                if (mapto_name not in self.been_mapto) and (mapto_name in self.patch_key_value.keys()):
                    self.been_mapto[mapto_name] = self.patch_key_value.get(mapto_name)
        for name in self.been_mapto:
            self.patch_key_value.pop(name)

    def boot_type_load_default(self):  # boot type无法load_default
        name = "BootType"
        default = self.att_pd.loc[name, "DefaultValue"]
        logging.info(f"BootType load default to \"{default}\"")
        if self.set_bios_option(**{name: default}).result:
            logging.info("BootType load default successfully")

    def try_patch_all(self):
        except_list = _BootInvolved.MemPop + _BootInvolved.Exclusive + _BootInvolved.ReadOnly
        for i in except_list:
            try:
                self.patch_key_value.pop(i)
            except Exception as e:
                logging.error(f"Exclude attribute: {e} not in Attributes list")

        for k in self.att_pd.index:
            if k not in self.patch_key_value.keys():
                logging.debug(f"Items {k} will be test in single patch")
                continue
            default_value = self.att_pd.loc[k, "DefaultValue"]
            self.result.loc[k, "default_value"] = default_value
        logging.info(f"Try Patch {len(self.patch_key_value)} items that without dependency")

        while True:
            write = self.set_bios_option(**self.patch_key_value)
            logging.info(f"PATCH Status: {write.status}")
            logging.info(f"PATCH Message: {write.body}")
            if write.result:
                for key, value in self.patch_key_value.items():
                    if isinstance(value, int):
                        logging.debug(f'\"{key}\":{value},')
                    else:
                        logging.debug(f'\"{key}\":\"{value}\",')
                    self.result.loc[key, "patch_status"] = "pass"
                break
            fail_att = re.search(r"Attributes/(\S+?)\s", write.body)
            if fail_att:
                fail_att = fail_att.group(1)
            else:
                logging.info(f"[Error] Attributes not found: {write.body}")
                continue
            logging.info(f"PATCH Failed: remove \"{fail_att}\" and retry...")
            self.result.loc[fail_att, "patch_status"] = f"{write.body}"
            if self.patch_key_value.get(fail_att):
                self.patch_key_value.pop(fail_att)
            logging.info(f"Remain {len(self.patch_key_value)} items in testing...")

    def narrow_down_fail_items(self):
        """ 批量修改失败时，10个1组筛选出是哪部分选项导致无法修改成功 """
        all_keys = list(self.test_fail_key_value.keys())
        for index in range(len(all_keys)//10+1):
            logging.info("================================================")
            self.bios_load_default()
            _reboot_sut()
            test_keys = all_keys[0:10] if len(all_keys)>10 else all_keys
            test_kv = {k: self.test_fail_key_value[k] for k in test_keys}
            patch = self.set_bios_option(**test_kv)
            logging.info(f"[NarrowDown]: PATCH {patch.body}")
            if patch.result:
                for key in test_kv:
                    self.result.loc[key, "patch_status"] = "pass"
                _reboot_sut()
            if self.check_bios_option(**test_kv):
                logging.info("[NarrowDown]: pass")
                for key in test_kv:
                    self.result.loc[key, "check_status"] = "pass"
                    self.test_fail_key_value.pop(key)
            else:
                logging.info("[NarrowDown]: fail")
            for tk in test_keys:
                all_keys.remove(tk)

    def boot_order_key_value(self):
        boot_order_name = "BootTypeOrder"
        boot_order_menus = [att for att in self.att_pd.index if boot_order_name in att]
        boot_order_values = self.att_pd.loc[f"{boot_order_name}0", "Value"]
        order_key_value = {}
        for order in boot_order_menus:
            choose_value = random.choice(boot_order_values)
            order_key_value[order] = choose_value
            boot_order_values.remove(choose_value)
        return order_key_value

    def exclude_items_key_value(self):
        exclude_kv = {}
        for boot_os in _BootInvolved.BootOS:
            exclude_kv[boot_os] = self.get_non_default_value(boot_os)
        return exclude_kv

    def single_patch_test(self):
        self.been_mapto.update(self.test_fail_key_value)  # fail的选项单独测试一遍
        self.been_mapto.update(self.exclude_items_key_value())  #排除的选项放在最后测试
        logging.info(f"{len(self.been_mapto)} items excluded, try test one by one...")
        excluded_items = deepcopy(self.been_mapto)
        for key, value in excluded_items.items():
            if key not in self.been_mapto:
                continue
            logging.info("============================================================")
            index_option = re.findall(r"(\w+)\[(\d+)]", key)
            patch_kv = {key: value}
            if index_option:  # 带索引的选项统一修改，减少重启次数
                index_name = index_option[0][0]
                if key in self.been_mapto.keys():
                    ifound = re.findall(rf"({index_name})\[(\d+)]", "".join(self.been_mapto.keys()))
                    for ikv in ifound:
                        ikey = f"{ikv[0]}[{ikv[1]}]"
                        patch_kv[ikey] = value
                        self.been_mapto.pop(ikey)
                else:
                    continue
            boot_order = re.findall(f"BootTypeOrder", key)
            if boot_order:  # BootOrder需要4个选项一起修改
                if key in self.been_mapto.keys():
                    patch_kv = self.boot_order_key_value()
                    for jkv in patch_kv:
                        self.been_mapto.pop(jkv)
                else:
                    continue
            logging.info(f"Testing: {patch_kv}")
            self.result.loc[key, "default_value"] = self.att_pd.loc[key, "DefaultValue"]
            self.bios_load_default()
            _reboot_sut()
            patch = self.set_bios_option(**patch_kv)
            if patch.result:
                logging.info(f"[PATCH] {patch.body} successfully")
                _reboot_sut()
                check_value = self.read_bios_option(*patch_kv.keys())
                for check_name in patch_kv.keys():
                    self.result.loc[check_name, "patch_status"] = "pass"
                    if check_value.get(check_name) == value:
                        self.result.loc[check_name, "check_status"] = "pass"
                        logging.info(f"[Check] {check_name} = {value} <pass>")
                    else:
                        logging.info(f"[Check] {check_name} = {value} <fail> -> {check_value.get(check_name)}")
                        self.result.loc[check_name, "check_status"] = "fail"
                continue
            logging.info(f"PATCH Status: {patch.status}")
            logging.info(f"PATCH Message: {patch.body}")
            for pf_name in patch_kv.keys():
                self.result.loc[pf_name, "patch_status"] = "fail"
                self.result.loc[pf_name, "message"] = patch.body

    def gen_report(self, name):
        self.result["summary"] = self.result.apply(
            lambda x:
            "Pass" if (
                    x.patch_status == "pass" and x.check_status == "pass") else
            ("Fail" if (
                    x.patch_status == "fail" or x.check_status == "fail") else
             None), axis=1)
        _to_excel(self.result, name)

    def run_test(self):
        self.bios_load_default()  # 测试前先恢复默认
        _reboot_sut()
        self.attributes_data()  # 获取所有变量信息
        self.gen_patch_key_value()  # 生成非默认值数据表
        self.report_init()
        self.drop_been_mapped()  # 去掉默认情况就存在联动关系的选项
        self.try_patch_all()  # 先尝试批量修改
        _reboot_sut()
        data = self.read_bios_option(*self.patch_key_value.keys())
        self.test_fail_key_value = {}
        for key, value in self.patch_key_value.items():  # 检查批量修改的结果
            if data.get(key) != value:
                self.test_fail_key_value[key] = value
                self.result.loc[key, "check_status"] = "fail"
                logging.info(rf"[Failed] {key} | patch=<{value}>; read=<{data.get(key)}>")
                continue
            self.result.loc[key, "check_status"] = "pass"
        self.gen_report("非联动测试结果批量修改")  # 批量修改结果生成报告
        self.boot_type_load_default()  # 恢复 BootType为UEFI（因为POST无法恢复）
        self.bios_load_default()
        _reboot_sut()
        self.narrow_down_fail_items()  # 测试Fail的10个1组，尝试PATCH，并更新测试结果数据表
        self.single_patch_test()  # 剩下的Fail选项 + 非默认情况下有联动关系的单独测试， 每次PATCH一项并重启，更新测试结果数据表
        # 测完恢复默认
        self.bios_load_default()
        _reboot_sut()
        # 生成测试报告
        self.gen_report("非联动测试结果汇总")


def _unitool_check(option, value):
    value = int(str(value), 16) if "0x" in str(value) else int(value)
    stdin, stdout, stderr = Sut.OS_SSH.ssh_client.exec_command(r'cd {};./unitool -r {} |grep value'.format(SutConfig.Env.UNI_PATH, option))
    rcv_data = stdout.read().decode("utf-8").split(":")
    if len(rcv_data) != 2:
        return False, rcv_data
    get_value = int(rcv_data[1])
    return (get_value == value), get_value


@core.test_case(("124", "[TC124] Redfish default value test", "Redfish default value test"))
def redfish_default_value_test():
    """
    检查Redfish的默认值和Unitool读到的默认值是否一致
    1， unitool读取默认值
    2， redfish读registry的默认值
    3， 比较二者是否一样
    """
    try:
        var_data = Sut.BMC_RFISH.Attributes()
        for index, att in enumerate(var_data):
            if var_data[index].get("Value"):
                var_data[index]["Value"] = [val["ValueName"] for val in att["Value"]]
        var_pd = pandas.DataFrame(var_data).set_index("AttributeName")
        redfish_info = [n.get("AttributeName") for n in Sut.BMC_RFISH.Attributes()]  # Redfish读到的信息
        BmcLib.force_reset()
        MiscLib.ping_sut(SutConfig.Env.OS_IP, 300)
        baseline_info = _option_value_index_map(base_xlsx=base_line)  # 基线文件定义的信息
        unitool_read_info = _read_default_values(redfish_info)  # Unitool读到的实际值
        default_value = {}
        fail_items = []
        for var_name in redfish_info:
            uni_name = "".join(re.findall("(.*?)\[", var_name)) if ("[0]" in var_name) else var_name
            if var_pd.loc[var_name, "Type"] == "Enumeration":
                baseline_var_name = baseline_info.get(uni_name)
                if isinstance(baseline_var_name, dict):
                    rfish_default = str(var_pd.loc[var_name, "DefaultValue"])
                    uni_rd_val_int = unitool_read_info.get(uni_name)
                    uni_read_val_name = baseline_var_name.get(uni_rd_val_int)
                    default_value[uni_name] = uni_read_val_name
                    if uni_read_val_name != rfish_default:
                        fail_items.append(
                            {"Variable": var_name, "Registry": rfish_default, "UnitoolRead": uni_read_val_name})
                        logging.info(
                            f'[{var_name}] Default: registry="{rfish_default}" | read="{uni_read_val_name}" | dic -> {baseline_var_name} | {uni_rd_val_int} | {uni_read_val_name}"')
            if var_pd.loc[var_name, "Type"] == "Integer":  # 数值型的变量基线中未指定默认值，暂时不做检查
                default_value[uni_name] = unitool_read_info.get(uni_name)
        if not fail_items:
            return core.Status.Pass
        return core.Status.Fail
    except Exception as e:
        logging.error(e)
        return core.Status.Fail


@core.test_case(("125", "[TC125] Redfish post load default test", "Redfish post load default test"))
def redfish_post_load_default_test():
    """
    1， 修改几个变量的值
    2， 重启检查是否改成功
    3， redfish POST恢复默认
    4， 重启检查是否全部恢复默认
    """

    pick_setup = {
        "UsbBoot": 0,
        "WakeOnPME": 1,
        "AcpiApicPolicy": 0,
        "FDMSupport": 0,
        "SataPort": 0,
        "sSataPort": 0,
        "PerformanceTuningMode": 0,
        "VTdSupport": 0,
        "ADDDCEn": 1,
        "ActiveCpuCores": 4,
        "ProcessorHyperThreadingDisable": 1,
        "UFSDisable": 1,
        "ProcessorEistEnable": 0,
        "C6Enable": 1,
        "IrqThreshold": 0,
        "EnableBiosSsaRMT": 1,
        "pprType": 0,
        "BMCWDTEnable": 1,
    }

    try:
        BmcLib.force_reset()
        MiscLib.ping_sut(SutConfig.Env.OS_IP, 300)
        pick_setup_default = deepcopy(pick_setup)
        val_default = Sut.UNITOOL.read(*pick_setup)  # read default
        for option in pick_setup:
            pick_setup_default[option] = val_default.get(option)
        Sut.UNITOOL.write(**pick_setup)  # config modify
        BmcLib.force_reset()
        MiscLib.ping_sut(SutConfig.Env.OS_IP, 300)
        val_modify = Sut.UNITOOL.read(*pick_setup)  # check modify
        fail_count = 0
        for option in pick_setup:
            if val_modify.get(option) != str(pick_setup.get(option)):
                logging.info(f"{option} set config failed")
                fail_count += 1
        Sut.BMC_RFISH.bios_load_default()
        BmcLib.force_reset()
        MiscLib.ping_sut(SutConfig.Env.OS_IP, 300)
        val_after_post = Sut.UNITOOL.read(*pick_setup)
        for option in pick_setup_default:
            if val_after_post.get(option) != pick_setup_default.get(option):
                logging.info(f"{option} don't load default after post performed")
                fail_count += 1
        if fail_count:
            return core.Status.Fail
        return core.Status.Pass
    except Exception as e:
        logging.error(e)
        return core.Status.Pass


# 【测试配置】
# 1. 配置 config 文件
#     -   bmc_ip
#     -   bmc_user
#     -   bmc_pw
#     -   COM
#     -   os_ip
#     -   os_user
#     -   os_pw
#     -   unitool_path
#     -   AppExcel
# 2. AppExcel变量名那一列的名字需要修改为 AttributeName
# 3. 测试完成后Excel报告保存在测试日志路径中
@core.test_case(("126", "[TC126] Redfish power efficiency test", "Redfish power efficiency test"))
def redfish_power_efficiency_test():
    """
    1. 按照能效菜单的Excel表格遍历修改能效菜单模式
    2. 重启进OS后，使用Unitool检查所有相关选项的值，确认是否与预期相同，相同则PASS，不相同则返回实际当前值
    4. Excel表格遍历完成后生成测试报告，以Excel格式保存
    """

    # HY5和2288V6的菜单名不一样
    app_name_list = ["ApplicationProfile", "BenchMarkSelection"]
    try:
        # os_ssh = SshConnection(config.os_ip, config.os_user, config.os_pw)
        pd = pandas.read_excel(power_efficiency, sheet_name=0, index_col="AttributeName")
        # rfish = Redfish(config.bmc_ip, config.bmc_user, config.bmc_pw)
        Sut.BMC_RFISH.registry_dump(True, path=SutConfig.Env.LOG_DIR, name="registry.json")  # 获取registry数据
        names = [n.get("AttributeName") for n in Sut.BMC_RFISH.Attributes()]
        app_name = "".join([app for app in app_name_list if app in names])

        for ap in pd:
            logging.info("===============================================")
            logging.info('[Set] {} = "{}"'.format(app_name, ap))
            pd[ap + "_Check"] = ""

            # PATCH AppProfile选项
            if not Sut.BMC_RFISH.set_bios_option(**{app_name: ap}).result:
                logging.info('Error: {} = {} PATCH Fail!'.format(app_name, repr(ap)))
                continue
            logging.info('{} = {} PATCH Pass!'.format(app_name, repr(ap)))

            # 重启后确认能效菜单是否为预期，并保存一份json文件
            _reboot_sut()
            if not Sut.BMC_RFISH.check_bios_option(**{app_name: ap}):
                logging.info('Current AppValue is "{}"'.format(Sut.BMC_RFISH.read_bios_option(app_name)))
                logging.info('Redfish Check: {} = {}  Fail'.format(app_name, repr(ap)))
                continue
            Sut.BMC_RFISH.current_dump(True, path=SutConfig.Env.LOG_DIR, name="{}.json".format(ap.replace("/", "")))
            logging.info('Redfish Check: {} = {}  Pass!'.format(app_name, repr(ap)))

            # 进入OS，unitool检查子选项的值
            if not Sut.OS_SSH.login():
                continue
            Sut.OS_SSH.ssh_client.exec_command(r'cd {};insmod ufudev.ko'.format(SutConfig.Env.UNI_PATH))
            logging.info("Unitool driver installed")
            time.sleep(3)  # 进OS后等待3秒确保Unitool Driver加载完成

            for sub in pd[ap].index:
                result, get_val = _unitool_check(sub, pd.loc[sub, ap])
                if not result:
                    pd.loc[sub, ap + "_Check"] = get_val
                    logging.info("[Check] {} = {} | fail ====-> {}".format(sub, pd.loc[sub, ap], get_val))
                    continue
                pd.loc[sub, ap + "_Check"] = "pass"
                logging.info("[Check] {} = {} | pass".format(sub, get_val))
            pd = pd.sort_index(axis=1)
        _to_excel(pd, "{}_Redfish_Test.xlsx".format(app_name))
        return core.Status.Pass
    except Exception as e:
        logging.error(e)
        return core.Status.Fail


@core.test_case(("127", "[TC127] Redfish dependency variables test", "Redfish dependency variables test"))
def redfish_dependency_test():
    """联动菜单测试"""
    try:
        deptest = _DepenTest()
        deptest.run_test()
        return core.Status.Pass
    except Exception as e:
        logging.error(e)
        return core.Status.Fail


@core.test_case(("128", "[TC128] Redfish non-dependency variables test", "Redfish non-dependency variables test"))
def redfish_non_dependency_test():
    """非联动菜单测试"""
    try:
        deptest = _NonDepTest()
        deptest.run_test()
        return core.Status.Pass
    except Exception as e:
        logging.error(e)
        return core.Status.Fail

