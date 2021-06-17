# -*- encoding=utf8 -*-
import os
import re
import time
import logging
import subprocess
import openpyxl
import pandas
from Core import SerialLib
import RedFish.config as config


def loginfo(preinfo=None, postinfo=None):
    def func_decorator(func):
        def func_wrapper(*args, **kwargs):
            if isinstance(preinfo, str):
                logging.info(repr(preinfo))
            rtn = func(*args, **kwargs)
            if isinstance(postinfo, str):
                logging.info(repr(postinfo))
            return rtn
        return func_wrapper
    return func_decorator


def reboot_to_os(ssh_bmc, timeout=600):
    cmd_off = 'ipmcset -d powerstate -v 2\n'
    cmd_on = 'ipmcset -d powerstate -v 1\n'
    ret_ensure = 'Do you want to continue'
    cmd_confirm = 'Y\n'
    ret_confirm = 'successfully'
    if not ssh_bmc.login():
        logging.info("BMC login fail")
        return
    logging.info('Reboot SUT...')
    if not ssh_bmc.interaction([cmd_off, cmd_confirm, cmd_on, cmd_confirm], [ret_ensure, ret_confirm, ret_ensure, ret_confirm]):
        logging.info("Reboot SUT Failed")
        return
    logging.info("Reboot SUT successful, wait sut online...")
    ssh_bmc.close_session()
    time.sleep(30)
    return True if ping_sut(timeout=timeout) else False


def reboot_sut(ssh_bmc, serial, msg="BIOS boot completed", timeout=config.os_timeout):
    cmd_off = 'ipmcset -d powerstate -v 2\n'
    cmd_on = 'ipmcset -d powerstate -v 1\n'
    ret_ensure = 'Do you want to continue'
    cmd_confirm = 'Y\n'
    ret_confirm = 'successfully'
    if not ssh_bmc.login():
        logging.info("BMC login fail")
        return
    logging.info('Reboot SUT...')
    if not ssh_bmc.interaction([cmd_off, cmd_confirm, cmd_on, cmd_confirm], [ret_ensure, ret_confirm, ret_ensure, ret_confirm]):
        logging.info("Reboot SUT Failed")
        return
    logging.info("Reboot SUT successful, wait sut boot completed...")
    ssh_bmc.close_session()
    if SerialLib.is_msg_present(serial, msg=msg, delay=timeout):
        SerialLib.clean_buffer(serial)
        return True


def ping_sut(ip=config.os_ip, timeout=180):
    logging.info("Ping OS IP Address: {} ...".format(ip))
    ping_cmd = 'ping {0}'.format(config.os_ip)
    start_time = time.time()
    while True:
        p = subprocess.Popen(args=ping_cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        (stdoutput, erroutput) = p.communicate()
        now = time.time()
        time_spent = (now - start_time)
        if 'TTL=' in stdoutput.decode('gbk'):
            logging.info("SUT is Online Now !")
            return True
        if time_spent > timeout:
            logging.info("Lost SUT for {} seconds, please check the OS ip".format(time_spent))
            return False


# 支持列表，字典和DataFrame转成excel文件
def to_excel(data, name="", path=os.path.abspath(config.REPORT_DIR)):
    if not os.path.exists(path):
        os.makedirs(path)
    now = time.strftime("%Y%m%d_%H%M%S", time.localtime())
    path_name = os.path.join(path, rf"{name} {now}.xlsx")
    if isinstance(data, list):
        pandas.Series(data).to_excel(path_name)
    elif isinstance(data, dict):
        pandas.DataFrame.from_dict(data, orient='index').to_excel(path_name)
    elif isinstance(data, pandas.DataFrame):
        data.to_excel(path_name)
    logging.info("Create excel {} successful!".format(path_name))


def read_default_values(unitool, variables):
    var_format = []
    for var in variables:
        if "[0]" in var:
            var_format.append(var[:var.find("[")])
        else:
            var_format.append(var)
    key_values = unitool.read(*var_format)
    return key_values


def option_value_index_map(base_xlsx):
    base_xlsx = openpyxl.load_workbook(base_xlsx)
    sheets = base_xlsx.sheetnames
    base_sheet = base_xlsx[sheets[0]]
    setup_col = "K"  # setup column index of xlsx
    value_col = "I"  # value column index of xlsx
    max_row = base_sheet.max_row
    table = {}
    for row in range(1, max_row):
        option = base_sheet[f"{setup_col}{row}"].value
        if not option:
            continue
        option = option.replace("\n", "").strip()
        values = base_sheet[f"{value_col}{row}"].value
        if not values:
            continue
        value_list = values.split("\n")
        for v in value_list:
            if len(re.findall(":", v)) == 1:
                vname = v.split(":")[0]
                vidx = v.split(":")[1].strip()
                index = str(int(vidx, 16)) if "0x" in vidx else vidx
                if isinstance(table.get(option), dict):
                    table[option].update({index: vname.strip()})
                else:
                    table[option] = {index: vname.strip()}
            elif len(re.findall(":", v)) > 1:
                vname = v[:v.rfind(":")]
                vidx = v[v.rfind(":")+1:]
                index = str(int(vidx, 16)) if "0x" in vidx else vidx
                if isinstance(table.get(option), dict) and table.get(option):
                    table[option].update({index: vname.strip()})
                else:
                    table[option] = {index: vname.strip()}
            else:
                table[option] = v
    return table
